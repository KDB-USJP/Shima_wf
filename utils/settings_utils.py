import json
import os
from pathlib import Path

from typing import Dict, Any, List, Optional, Set, Union, cast

class ShimaSettings:
    """Utility to load and cache centralized site settings with fallbacks."""
    
    _config: Optional[Dict[str, Any]] = None
    _extension_root: Path = Path(__file__).parent.parent
    _config_path: Path = _extension_root / "config" / "site_default_settings.json"
    _user_config: Optional[Dict[str, Any]] = None
    _user_config_path: Path = Path(__file__).parent.parent / "config" / "shima_settings.json"
    _excel_palettes: Optional[Dict[str, Any]] = None

    # --- FALLBACKS ---
    DEFAULT_API_BASE = "https://shima.wf"
    DEFAULT_STYLETHUMBS = {
        "walking_woman": "https://www.shima.wf/api/assets/download?pack=walking_woman",
        "still_life_classic": "https://www.shima.wf/api/assets/download?pack=still_life_classic"
    }

    DEFAULT_COMMONS = {
        "model_types": [
            "sdxl", "sd1.5", "sd2.x", "sd3",
            "flux", "pony", "illustrious",
            "auraflow", "hunyuan",
            "lumina2", "chroma", "hidream",
            "z-image-base", "z-image-turbo"
        ],
        "aspect_ratios": [
            "1:1 Square", 
            "16:9 Widescreen", 
            "4:3 Standard", 
            "21:9 Ultrawide", 
            "3:2 Photo",
            "Custom"
        ],
        "orientations": ["landscape", "portrait", "auto"]
    }

    DEFAULT_MULTISAVER = {
        "filename_order_presets": [
            "PRE,PRJ,BN,ET,SUF,TS,CID",
            "BN,ET,TS,CID"
        ],
        "separators": ["_", "-", ".", " ", ""]
    }

    @classmethod
    def reload_excel_palettes(cls):
        """Force a hot-reload of the Excel theme sheet."""
        cls._excel_palettes = None
        return cls.get_excel_palettes()

    @classmethod
    def get_excel_palettes(cls):
        """Parse E:\\ComfyDev\\Shima\\assets\\data\\shima_sheets.xlsx (tab: node-color-themes) once."""
        if cls._excel_palettes is not None:
            return cls._excel_palettes
            
        # Dynamically resolve path relative to this file
        # __file__ is in utils/, so go up one level to Shima root, then into assets/data
        excel_path = Path(__file__).parent.parent / "assets" / "data" / "shima_sheets.xlsx"
        palettes: Dict[str, Any] = {}
        
        if not excel_path.exists():
            cls._excel_palettes = palettes
            return palettes

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            if "node-color-themes" not in wb.sheetnames:
                return palettes
            
            sheet = wb["node-color-themes"]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return palettes

            # Header row (starting col 2) has theme names
            header = rows[0]
            theme_names = [name for name in header[1:] if name]
            
            # Initialize palettes
            for name in theme_names:
                palettes[name] = {"node": {}}

            # Data rows
            data_rows = cast(List[Any], rows[1:])
            for row in data_rows:
                node_key = row[0]
                if not node_key:
                    continue
                
                for idx, color in enumerate(row[1:]):
                    if idx < len(theme_names) and color:
                        theme_name = theme_names[idx]
                        palettes[theme_name]["node"][node_key] = color

            cls._excel_palettes = palettes
            return palettes
        except Exception as e:
            print(f"[Shima] Error parsing Excel palettes: {e}")
            cls._excel_palettes = palettes
            return palettes

    @classmethod
    def get_config(cls):
        if cls._config is None:
            cls.reload()
        
        # Merge Excel palettes into config
        config = cls._config.copy()
        if "themes" not in config:
            config["themes"] = {}
        
        excel_palettes = cls.get_excel_palettes()
        if excel_palettes:
            if "palettes" not in config["themes"]:
                config["themes"]["palettes"] = {}
            config["themes"]["palettes"].update(excel_palettes)
            
        return config

    @classmethod
    def get_user_config(cls):
        # ... rest of the file ...
        if cls._user_config is None:
            cls.reload_user()
        return cls._user_config

    @classmethod
    def get_manifests_dir(cls) -> Path:
        """Returns the path to custom hub manifests."""
        path = Path(__file__).parent.parent / "custom_content" / "hub_manifests"
        path.mkdir(parents=True, exist_ok=True)
        return path

    @classmethod
    def load_manifest_config(cls, manifest_name: str) -> Dict[str, Any]:
        """Loads a manifest by name, merging it with site defaults."""
        import copy
        base_config = copy.deepcopy(cls.get_config())
        
        # Sanitize manifest_name (JS might send "undefined" or null as string)
        if not manifest_name or manifest_name in ["Default", "null", "undefined", "None"]:
            return base_config
            
        manifest_path = cls.get_manifests_dir() / manifest_name
        if not manifest_path.exists():
            # Try lower-case fallback for cross-platform robustness
            manifest_path = cls.get_manifests_dir() / manifest_name.lower()
            
        if not manifest_path.exists():
            print(f"[Shima] Warning: Manifest {manifest_name} not found at {manifest_path}")
            return base_config
            
        try:
            with open(manifest_path, "r") as f:
                custom_data = json.load(f)
                
            # --- REPLACE Strategy (Context Switch) ---
            # If bundles are present in manifest, we show ONLY those bundles
            if "bundles" in custom_data:
                base_config["bundles"] = custom_data["bundles"]
                
            # If models are present, we show ONLY those models
            if "curated_models" in custom_data:
                base_config["curated_models"] = custom_data["curated_models"]

            # If nodes are present, we show ONLY those nodes
            if "curated_nodes" in custom_data:
                base_config["curated_nodes"] = custom_data["curated_nodes"]

            # Add support for generic assets targeting input/shima_custom/[manifest]
            if "assets" in custom_data:
                base_config["assets"] = custom_data["assets"]
            else:
                base_config["assets"] = {}
            
            # Update registry if provided
            if "node_registry" in custom_data:
                if "node_registry" not in base_config: base_config["node_registry"] = {}
                base_config["node_registry"].update(custom_data["node_registry"])
                
            return base_config
        except Exception as e:
            print(f"[Shima] Error loading manifest {manifest_name}: {e}")
            return base_config

    @classmethod
    def reload(cls):
        if cls._config_path.exists():
            try:
                with open(cls._config_path, "r") as f:
                    cls._config = json.load(f)
                    print(f"[Shima] Loaded site settings from {cls._config_path}")
            except Exception as e:
                print(f"[Shima] Error loading site settings: {e}")
                cls._config = {}
        else:
            cls._config = {}

        # Ensure YourStyleImages directory exists
        style_dir = cls._extension_root / "assets" / "styles" / "YourStyleImages"
        style_dir.mkdir(parents=True, exist_ok=True)

    @classmethod
    def reload_user(cls):
        if cls._user_config_path.exists():
            try:
                with open(cls._user_config_path, "r") as f:
                    cls._user_config = json.load(f)
                    print(f"[Shima] Loaded user settings from {cls._user_config_path}")
            except Exception as e:
                print(f"[Shima] Error loading user settings: {e}")
                cls._user_config = {}
        else:
            cls._user_config = {}

    @classmethod
    def get_api_base(cls):
        """Returns the server URL from user settings, falling back to default."""
        return cls.get_user_config().get("api_base", cls.DEFAULT_API_BASE)

    @classmethod
    def get_asset_packs(cls):
        """Returns the list of style packs, ensuring 'YourStyleImages' is always included."""
        packs = cls.get_config().get("stylethumbs", cls.DEFAULT_STYLETHUMBS)
        
        # Ensure YourStyleImages is in the list (as a local-only entry)
        if "YourStyleImages" not in packs:
            packs["YourStyleImages"] = "local"
            
        return packs

    @classmethod
    def get_commons(cls):
        return cls.get_config().get("commons", cls.DEFAULT_COMMONS)

    @classmethod
    def get_multisaver(cls):
        return cls.get_config().get("multisaver", cls.DEFAULT_MULTISAVER)

    @classmethod
    def get_bundles(cls):
        return cls.get_config().get("bundles", {})

    @classmethod
    def get_node_registry(cls):
        return cls.get_config().get("node_registry", {})

    @classmethod
    def get_curated_nodes(cls):
        return cls.get_config().get("curated_nodes", [])

    @classmethod
    def get_curated_models(cls):
        return cls.get_config().get("curated_models", [])

    @classmethod
    def get_credential_fallbacks(cls):
        return cls.get_config().get("credential_fallbacks", {})

    @classmethod
    def get_civitai_key(cls):
        return cls.get_user_config().get("civitai_key", "")

    @classmethod
    def get_hf_token(cls):
        return cls.get_user_config().get("hf_token", "")

    @classmethod
    def get_list(cls, section, key, default):
        """Helper to safely get a list from a section."""
        return cls.get_config().get(section, {}).get(key, default)
